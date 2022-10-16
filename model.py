import cv2
import numpy as np
import os
import sys
import tensorflow as tf
import openpyxl

EPOCHS = 8
IMG_WIDTH = 28
IMG_HEIGHT = 28
NUM_CATEGORIES = 25
TEST_DATA_PATH = "data\\sign_mnist_test.xlsx"
TRAIN_DATA_PATH = "data\\sign_mnist_train.xlsx"


def main():
    # Get image arrays and labels for training
    training_images, training_labels = load_data(TRAIN_DATA_PATH)

    # Split data into training and testing sets
    training_labels = tf.keras.utils.to_categorical(training_labels, NUM_CATEGORIES).astype(np.uint8)

    # Get the model
    model = get_model()

    # Fit model on training data
    model.fit(np.array(training_images), np.array(training_labels), epochs=EPOCHS)

    # Get image arrays and labels for testing
    testing_images, testing_labels = load_data(TEST_DATA_PATH)

    testing_labels = tf.keras.utils.to_categorical(testing_labels)

    # Evaluate neural network performance
    model.evaluate(testing_images,  testing_labels, verbose=2)

    # Save model to file
    if len(sys.argv) == 2:
        filename = sys.argv[1]
        model.save(filename)
        print(f"Model saved to {filename}.")


def get_model():
    """
    Returns a compiled model. Assume that the
    'input_shape' is (IMG_WIDTH, IMG_HEIGHT, 1)
    The output layer should have `NUM_CATEGORIES`
    """

    model = tf.keras.models.Sequential(
        [   
            #Convolutional layer
            tf.keras.layers.Conv2D(32, (2, 2), input_shape=(IMG_WIDTH, IMG_HEIGHT, 1), activation = "relu"),
            tf.keras.layers.Conv2D(64, (2, 2), activation = "relu"),
            #Poolin
            tf.keras.layers.MaxPooling2D(pool_size=(2, 2)),
            tf.keras.layers.Dropout(.25),


            #Convolutional layer
            tf.keras.layers.Conv2D(32, (2, 2), activation = "relu"),
            tf.keras.layers.Conv2D(64, (2, 2), activation = "relu"),
            #Poolin
            tf.keras.layers.MaxPooling2D(pool_size=(2, 2)),
            tf.keras.layers.Dropout(.25),

            #Flat as a pancake
            tf.keras.layers.Flatten(),

            #Hidden layer with 128 nodes, ReLU activation
            tf.keras.layers.Dense(1024, activation = "relu"),
            tf.keras.layers.Dropout(.5),
            
            #Output layer
            tf.keras.layers.Dense(NUM_CATEGORIES, activation = "softmax")
        ]
    )

    #Compile and return the model
    model.compile(optimizer = tf.keras.optimizers.Adam(learning_rate = 0.001, decay = 0.0005), loss = "categorical_crossentropy", metrics = ["accuracy"])
    return model


def load_data(data_dir):
    """Load image data from directory 'data_dir'.

    Return tuple '(images, labels)'. 'images' should be a list of all
    of the images in the data directory, where each image is formatted as a
    numpy ndarray with dimensions IMG_WIDTH x IMG_HEIGHT. 'labels' should
    be a list of integer labels, representing the categories for each of the
    corresponding 'images'.
    """
    images = []
    labels = []
    # numrows, numcols
    from openpyxl import load_workbook
    wb_obj = load_workbook(data_dir)
    sheet_obj = wb_obj.active
    numrows = sheet_obj.max_row
    numcolumns = sheet_obj.max_column
    # note: the row and column numbers start at 1
    os.system('cls')
    #have two for loops, one that sto,res the label number, and the other 
    #that parses through the row
    for i in range(2, numrows + 1):
        labels.append(sheet_obj.cell(row=i, column=1).value)
        image = []
        for u in range (2, numcolumns + 1):
            image.append(sheet_obj.cell(row=i, column=u).value)
        images.append(np.array(image))
    return (np.array(images).reshape((-1, 28, 28, 1)).astype(np.uint8), labels)


if __name__ == "__main__":
    main()


